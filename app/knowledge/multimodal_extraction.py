"""Local, project-scoped extraction for registered evidence assets."""

from __future__ import annotations

import csv
from hashlib import sha256
import importlib.util
import json
import os
from pathlib import Path, PurePosixPath
import re
import shutil
import subprocess
from typing import Any

from app.knowledge.extraction_reference_projection import ExtractionReferenceProjector
from app.knowledge.wiki_contracts import ExtractionArtifact, ExtractionStatus, TableArtifact
from app.knowledge.wiki_repository import WikiRepository


_MAX_ASSET_BYTES = 5 * 1024 * 1024
_MAX_TABLE_ROWS = 10_000
_MAX_TABLE_COLUMNS = 100
_MAX_DERIVATIVE_CHARS = 500_000
_MAX_MEDIA_DURATION_SECONDS = 7 * 24 * 60 * 60
CURRENT_EXTRACTOR_REVISION = "local-v2"
_LZSTRING_BASE64_ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/="
_LZSTRING_BASE64_VALUES = {value: index for index, value in enumerate(_LZSTRING_BASE64_ALPHABET)}


def _decompress_lzstring_base64(value: str) -> str | None:
    """Decode Excalidraw's whitespace-wrapped LZString Base64 scene data."""
    compressed = "".join(value.split())
    if not compressed:
        return None

    try:
        length = len(compressed)

        def next_value(index: int) -> int:
            if index >= length:
                raise ValueError("compressed scene ended unexpectedly")
            return _LZSTRING_BASE64_VALUES[compressed[index]]

        data_value = next_value(0)
        data_position = 32
        data_index = 1

        def read_bits(width: int) -> int:
            nonlocal data_value, data_position, data_index
            bits = 0
            power = 1
            limit = 1 << width
            while power != limit:
                if data_value & data_position:
                    bits |= power
                data_position >>= 1
                if data_position == 0:
                    data_position = 32
                    data_value = next_value(data_index)
                    data_index += 1
                power <<= 1
            return bits

        dictionary: dict[int, str | int] = {0: 0, 1: 1, 2: 2}
        enlarge_in = 4
        dictionary_size = 4
        bit_width = 3
        initial = read_bits(2)
        if initial == 0:
            current = chr(read_bits(8))
        elif initial == 1:
            current = chr(read_bits(16))
        elif initial == 2:
            return ""
        else:
            return None

        dictionary[3] = current
        previous = current
        output = [current]
        while True:
            if data_index > length:
                return None
            code = read_bits(bit_width)
            if code == 0:
                dictionary[dictionary_size] = chr(read_bits(8))
                dictionary_size += 1
                code = dictionary_size - 1
                enlarge_in -= 1
            elif code == 1:
                dictionary[dictionary_size] = chr(read_bits(16))
                dictionary_size += 1
                code = dictionary_size - 1
                enlarge_in -= 1
            elif code == 2:
                return "".join(output)

            if enlarge_in == 0:
                enlarge_in = 1 << bit_width
                bit_width += 1

            candidate = dictionary.get(code)
            if isinstance(candidate, str):
                entry = candidate
            elif code == dictionary_size:
                entry = previous + previous[0]
            else:
                return None

            output.append(entry)
            dictionary[dictionary_size] = previous + entry[0]
            dictionary_size += 1
            enlarge_in -= 1
            previous = entry
            if enlarge_in == 0:
                enlarge_in = 1 << bit_width
                bit_width += 1
    except (KeyError, ValueError):
        return None



class LocalMultimodalExtractor:
    """Create bounded derivatives without altering Vault originals or source bodies."""

    def __init__(self, repository: WikiRepository, vault_root: Path | str) -> None:
        self.repository = repository
        self.vault_root = Path(vault_root).resolve()
        if not self.vault_root.is_dir():
            raise ValueError("Obsidian Vault root does not exist")
        self.reference_projector = ExtractionReferenceProjector(repository)

    @staticmethod
    def capabilities() -> dict[str, dict[str, str]]:
        """Report only locally observable capabilities, never inferred success."""
        module_available = lambda name: importlib.util.find_spec(name) is not None
        tesseract = LocalMultimodalExtractor._binary("tesseract")
        ffprobe = LocalMultimodalExtractor._binary("ffprobe")
        return {
            "csv_table": {"state": "available", "detail": "python_stdlib"},
            "canvas": {"state": "available", "detail": "json_parser"},
            "pdf_text": {"state": "available" if module_available("pdfplumber") else "unavailable", "detail": "pdfplumber"},
            "spreadsheet": {"state": "available" if module_available("openpyxl") else "unavailable", "detail": "openpyxl"},
            "image_metadata": {"state": "available" if module_available("PIL") else "unavailable", "detail": "Pillow"},
            "ocr": {"state": "available" if tesseract else "unavailable", "detail": "tesseract"},
            "scanned_pdf_ocr": {
                "state": "available" if tesseract and module_available("fitz") else "unavailable",
                "detail": "tesseract+pymupdf",
            },
            "media_probe": {"state": "available" if ffprobe else "unavailable", "detail": "ffprobe"},
        }

    @staticmethod
    def _binary(name: str) -> str | None:
        """Resolve an extractor binary without assuming the current process PATH is fresh."""
        configured = {
            "tesseract": os.environ.get("KNOWLEDGE_TESSERACT_PATH", ""),
            "ffprobe": os.environ.get("KNOWLEDGE_FFPROBE_PATH", ""),
        }.get(name, "").strip()
        if configured and Path(configured).is_file():
            return configured

        discovered = shutil.which(name)
        if discovered:
            return discovered

        if os.name == "nt" and name == "tesseract":
            for root_name in ("ProgramFiles", "ProgramW6432", "ProgramFiles(x86)"):
                root = os.environ.get(root_name, "").strip()
                candidate = Path(root) / "Tesseract-OCR" / "tesseract.exe" if root else None
                if candidate and candidate.is_file():
                    return str(candidate)
        if os.name == "nt" and name == "ffprobe":
            local_app_data = os.environ.get("LOCALAPPDATA", "").strip()
            packages = Path(local_app_data) / "Microsoft" / "WinGet" / "Packages" if local_app_data else None
            if packages and packages.is_dir():
                for candidate in sorted(packages.glob("Gyan.FFmpeg_*/ffmpeg-*/bin/ffprobe.exe"), reverse=True):
                    if candidate.is_file():
                        return str(candidate)
        return None

    def extract(
        self,
        *,
        project_id: str,
        source_id: str,
        asset_id: str,
        extractor_revision: str = CURRENT_EXTRACTOR_REVISION,
    ) -> dict[str, Any]:
        asset = self.repository.get_media_asset(project_id, asset_id)
        if not asset or asset["source_id"] != source_id:
            raise KeyError("asset is missing or belongs to another source")
        path, resolution_error = self._resolve(asset["storage_ref"])
        if resolution_error:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="local-asset-access",
                extractor_revision=extractor_revision,
                input_hash=asset["byte_hash"],
                status=ExtractionStatus.RESTRICTED if resolution_error == "asset_path_restricted" else ExtractionStatus.UNAVAILABLE,
                error=resolution_error,
            )
        assert path is not None
        if path.stat().st_size > _MAX_ASSET_BYTES:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="local-asset-access",
                extractor_revision=extractor_revision,
                input_hash=asset["byte_hash"],
                status=ExtractionStatus.UNAVAILABLE,
                error="asset_exceeds_local_extraction_limit",
                metadata={"byte_size": path.stat().st_size, "limit_bytes": _MAX_ASSET_BYTES},
            )
        payload = path.read_bytes()
        input_hash = sha256(payload).hexdigest()
        if input_hash != asset["byte_hash"]:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="local-integrity-check",
                extractor_revision=extractor_revision,
                input_hash=input_hash,
                status=ExtractionStatus.NEEDS_REVIEW,
                error="original_hash_changed",
            )

        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix in {".xlsx", ".xls"}:
            return self._extract_xlsx(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix in {".canvas", ".excalidraw"}:
            return self._extract_canvas(project_id, source_id, asset_id, payload, extractor_revision)
        if path.name.lower().endswith(".excalidraw.md"):
            return self._extract_excalidraw_markdown(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix in {".txt", ".md"}:
            return self._extract_text(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix == ".pdf":
            return self._extract_pdf(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
            return self._extract_image(project_id, source_id, asset_id, payload, extractor_revision)
        if suffix in {".mp3", ".mp4", ".wav", ".m4a", ".mov"}:
            return self._extract_media_probe(project_id, source_id, asset_id, payload, path, extractor_revision)
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="local-file-type",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            status=ExtractionStatus.UNSUPPORTED,
            error="unsupported_asset_type",
            metadata={"extension": suffix},
        )

    def _extract_csv(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        try:
            text = payload.decode("utf-8-sig")
            rows = list(csv.reader(text.splitlines()))
        except (UnicodeDecodeError, csv.Error):
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="csv-table",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="csv_decode_failed",
            )
        if not rows:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="csv-table",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.NEEDS_REVIEW,
                error="csv_has_no_rows",
            )
        schema = [cell.strip()[:256] or f"column_{index + 1}" for index, cell in enumerate(rows[0][:_MAX_TABLE_COLUMNS])]
        data_rows = [row[:_MAX_TABLE_COLUMNS] for row in rows[1:_MAX_TABLE_ROWS + 1]]
        derivative = "\n".join("\t".join(cell[:1_024] for cell in row) for row in [schema, *data_rows])[:_MAX_DERIVATIVE_CHARS]
        artifact = self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="csv-table",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=derivative,
            status=ExtractionStatus.COMPLETE,
            metadata={"row_count": len(data_rows), "column_count": len(schema), "truncated": len(rows) - 1 > len(data_rows)},
        )
        self._record_table(artifact, project_id, source_id, schema, len(data_rows), extractor_revision)
        return artifact

    def _extract_xlsx(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        if self.capabilities()["spreadsheet"]["state"] != "available":
            return self._unavailable(project_id, source_id, asset_id, payload, "spreadsheet", extractor_revision)
        try:
            from io import BytesIO
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
            sheet = workbook.active
            rows = [
                ["" if value is None else str(value) for value in row[:_MAX_TABLE_COLUMNS]]
                for row in sheet.iter_rows(values_only=True)
            ][:_MAX_TABLE_ROWS + 1]
        except Exception:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="xlsx-table",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="spreadsheet_extraction_failed",
            )
        if not rows:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="xlsx-table",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.NEEDS_REVIEW,
                error="spreadsheet_has_no_rows",
            )
        schema = [cell.strip()[:256] or f"column_{index + 1}" for index, cell in enumerate(rows[0])]
        data_rows = rows[1:]
        derivative = "\n".join("\t".join(cell[:1_024] for cell in row) for row in [schema, *data_rows])[:_MAX_DERIVATIVE_CHARS]
        artifact = self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="xlsx-table",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=derivative,
            status=ExtractionStatus.COMPLETE,
            metadata={"row_count": len(data_rows), "column_count": len(schema), "sheet": sheet.title},
        )
        self._record_table(artifact, project_id, source_id, schema, len(data_rows), extractor_revision)
        return artifact

    def _extract_canvas(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        try:
            document = json.loads(payload.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="canvas-elements",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="canvas_json_invalid",
            )
        nodes = document.get("nodes", []) if isinstance(document, dict) else []
        text_parts = [str(node.get("text") or "").strip() for node in nodes if isinstance(node, dict)]
        text = "\n".join(value for value in text_parts if value)[:_MAX_DERIVATIVE_CHARS]
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="canvas-elements",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=text,
            status=ExtractionStatus.COMPLETE if text else ExtractionStatus.PARTIAL,
            metadata={"node_count": len(nodes), "text_node_count": len([value for value in text_parts if value])},
        )

    def _extract_text(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        try:
            content = payload.decode("utf-8")[:_MAX_DERIVATIVE_CHARS]
        except UnicodeDecodeError:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="utf8-text",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="utf8_decode_failed",
            )
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="utf8-text",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=content,
            status=ExtractionStatus.COMPLETE,
        )

    def _extract_excalidraw_markdown(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        try:
            markdown = payload.decode("utf-8")
        except UnicodeDecodeError:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="excalidraw-elements",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="excalidraw_decode_failed",
            )
        document: dict[str, Any] | None = None
        scene_encoding = ""
        for match in re.finditer(r"```(?P<encoding>compressed-json|json)\s*\n(?P<payload>.*?)\n```", markdown, re.DOTALL):
            encoding = str(match.group("encoding"))
            raw_json = str(match.group("payload")).strip()
            if encoding == "compressed-json":
                if not raw_json:
                    document = {"elements": []}
                    scene_encoding = encoding
                    break
                raw_json = _decompress_lzstring_base64(raw_json) or ""
                if not raw_json:
                    return self._record(
                        project_id=project_id,
                        source_id=source_id,
                        asset_id=asset_id,
                        extractor="excalidraw-elements",
                        extractor_revision=extractor_revision,
                        input_hash=sha256(payload).hexdigest(),
                        status=ExtractionStatus.FAILED,
                        error="excalidraw_compressed_scene_decode_failed",
                        metadata={"scene_encoding": encoding},
                    )
            try:
                parsed = json.loads(raw_json)
            except json.JSONDecodeError:
                continue
            if isinstance(parsed, dict) and isinstance(parsed.get("elements"), list):
                document = parsed
                scene_encoding = encoding
                break
        elements = document.get("elements", []) if document else []
        text = "\n".join(
            str(element.get("text") or element.get("rawText") or "").strip()
            for element in elements
            if isinstance(element, dict)
        )[:_MAX_DERIVATIVE_CHARS]
        metadata = {
            "element_count": len(elements),
            "drawing_json_detected": document is not None,
            "scene_encoding": scene_encoding or "unknown",
        }
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="excalidraw-elements",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=text,
            status=ExtractionStatus.COMPLETE if text else ExtractionStatus.PARTIAL,
            error="" if text else "excalidraw_no_elements" if document is not None else "excalidraw_drawing_json_missing",
            metadata=metadata,
        )

    def _extract_image(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        if self.capabilities()["image_metadata"]["state"] != "available":
            return self._unavailable(project_id, source_id, asset_id, payload, "image_metadata", extractor_revision)
        try:
            from io import BytesIO
            from PIL import Image

            image = Image.open(BytesIO(payload))
            metadata = {"width": image.width, "height": image.height, "mode": image.mode, "format": image.format or ""}
        except Exception:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="image-metadata",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="image_metadata_extraction_failed",
            )
        ocr = self._ocr(payload)
        content = ocr[:_MAX_DERIVATIVE_CHARS]
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="image-ocr" if content else "image-metadata",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            content=content,
            status=ExtractionStatus.COMPLETE if content else ExtractionStatus.PARTIAL,
            error="" if content else "ocr_unavailable",
            metadata={**metadata, "ocr": "complete" if content else self.capabilities()["ocr"]},
        )

    def _extract_pdf(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        extractor_revision: str,
    ) -> dict[str, Any]:
        if self.capabilities()["pdf_text"]["state"] != "available":
            return self._unavailable(project_id, source_id, asset_id, payload, "pdf_text", extractor_revision)
        try:
            import pdfplumber
            from io import BytesIO

            with pdfplumber.open(BytesIO(payload)) as pdf:
                pages = [(page.extract_text() or "").strip() for page in pdf.pages]
            content = "\n\n".join(text for text in pages if text)[:_MAX_DERIVATIVE_CHARS]
        except Exception:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="pdf-text",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="pdf_text_extraction_failed",
            )
        if content:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="pdf-text",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                content=content,
                status=ExtractionStatus.COMPLETE,
                metadata={"page_count": len(pages)},
            )

        ocr_content = self._ocr_pdf(payload)
        if ocr_content:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="pdf-ocr",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                content=ocr_content,
                status=ExtractionStatus.COMPLETE,
                metadata={"page_count": len(pages), "ocr": "complete"},
            )

        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="pdf-text",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            status=ExtractionStatus.NEEDS_REVIEW,
            error="pdf_requires_ocr",
            metadata={"page_count": len(pages), "ocr": self.capabilities()["scanned_pdf_ocr"]},
        )

    def _ocr_pdf(self, payload: bytes) -> str:
        if self.capabilities()["scanned_pdf_ocr"]["state"] != "available":
            return ""
        try:
            import fitz

            document = fitz.open(stream=payload, filetype="pdf")
            try:
                pages = [
                    self._ocr(page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False).tobytes("png"))
                    for page in document
                ]
            finally:
                document.close()
        except Exception:
            return ""
        return "\n\n".join(text for text in pages if text)[:_MAX_DERIVATIVE_CHARS]

    def _extract_media_probe(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        path: Path,
        extractor_revision: str,
    ) -> dict[str, Any]:
        binary = self._binary("ffprobe")
        if not binary:
            return self._unavailable(project_id, source_id, asset_id, payload, "media_probe", extractor_revision)
        try:
            result = subprocess.run(
                [binary, "-v", "error", "-show_format", "-show_streams", "-of", "json", str(path)],
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                check=False,
                timeout=45,
            )
            probe = json.loads(result.stdout.decode("utf-8", errors="replace")) if result.returncode == 0 else {}
        except (OSError, subprocess.TimeoutExpired, json.JSONDecodeError):
            probe = {}
        if not isinstance(probe, dict):
            probe = {}
        if not probe:
            return self._record(
                project_id=project_id,
                source_id=source_id,
                asset_id=asset_id,
                extractor="ffprobe-media-metadata",
                extractor_revision=extractor_revision,
                input_hash=sha256(payload).hexdigest(),
                status=ExtractionStatus.FAILED,
                error="media_probe_failed",
            )
        streams = probe.get("streams") if isinstance(probe.get("streams"), list) else []
        format_data = probe.get("format") if isinstance(probe.get("format"), dict) else {}
        try:
            duration = float(format_data.get("duration"))
        except (TypeError, ValueError):
            duration = 0.0
        if duration < 0 or duration > _MAX_MEDIA_DURATION_SECONDS:
            duration = 0.0
        metadata = {
            "audio_streams": sum(1 for stream in streams if isinstance(stream, dict) and stream.get("codec_type") == "audio"),
            "container": str(format_data.get("format_name") or "")[:128],
            "duration_seconds": duration,
            "stream_count": len(streams),
            "video_streams": sum(1 for stream in streams if isinstance(stream, dict) and stream.get("codec_type") == "video"),
        }
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor="ffprobe-media-metadata",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            status=ExtractionStatus.PARTIAL,
            error="transcription_unavailable",
            metadata=metadata,
        )

    def _unavailable(
        self,
        project_id: str,
        source_id: str,
        asset_id: str,
        payload: bytes,
        capability: str,
        extractor_revision: str,
    ) -> dict[str, Any]:
        return self._record(
            project_id=project_id,
            source_id=source_id,
            asset_id=asset_id,
            extractor=f"local-{capability}",
            extractor_revision=extractor_revision,
            input_hash=sha256(payload).hexdigest(),
            status=ExtractionStatus.UNAVAILABLE,
            error=f"{capability}_unavailable",
            metadata={"capability": self.capabilities()[capability]},
        )

    def _record(self, **values: Any) -> dict[str, Any]:
        content = str(values.pop("content", ""))
        artifact = self.repository.create_extraction_artifact(
            ExtractionArtifact(
                **values,
                content=content,
                content_hash=sha256(content.encode("utf-8")).hexdigest() if content else "",
            )
        )
        self.reference_projector.project_extraction_id(str(artifact["project_id"]), str(artifact["id"]))
        return artifact

    def _record_table(
        self,
        artifact: dict[str, Any],
        project_id: str,
        source_id: str,
        schema: list[str],
        row_count: int,
        extractor_revision: str,
    ) -> None:
        content = self.repository.get_extraction_content(project_id, artifact["id"])
        content_hash = str(content.get("content_hash") or "") if content else ""
        if not content_hash:
            return
        table = self.repository.create_table_artifact(
            TableArtifact(
                project_id=project_id,
                source_id=source_id,
                extraction_id=artifact["id"],
                schema=schema,
                row_count=row_count,
                units=self._infer_units(schema),
                content_hash=content_hash,
                metadata={"extractor": artifact["extractor"], "extractor_revision": extractor_revision},
            )
        )
        self.reference_projector.project_table_id(project_id, str(table["id"]))

    def _ocr(self, payload: bytes) -> str:
        binary = self._binary("tesseract")
        if not binary:
            return ""
        try:
            result = subprocess.run(
                [binary, "stdin", "stdout"],
                input=payload,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                check=False,
                timeout=45,
            )
        except (OSError, subprocess.TimeoutExpired):
            return ""
        return result.stdout.decode("utf-8", errors="replace").strip() if result.returncode == 0 else ""

    def _resolve(self, storage_ref: str) -> tuple[Path | None, str]:
        candidate = self.vault_root.joinpath(*PurePosixPath(storage_ref).parts)
        try:
            resolved = candidate.resolve()
            resolved.relative_to(self.vault_root)
        except (OSError, ValueError):
            return None, "asset_path_restricted"
        if candidate.is_symlink() or not resolved.is_file():
            return None, "asset_missing"
        return resolved, ""

    @staticmethod
    def _infer_units(schema: list[str]) -> dict[str, str]:
        units: dict[str, str] = {}
        for value in schema:
            normalized = value.lower()
            if normalized.endswith("_usd"):
                units[value] = "USD"
            elif normalized.endswith("_percent") or normalized.endswith("_pct"):
                units[value] = "percent"
        return units
